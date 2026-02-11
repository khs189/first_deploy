from openpyxl import load_workbook

def load_sheet(path: str):
    wb = load_workbook(path)
    ws = wb.active
    return wb, ws

def collect_target_rows(ws, start_row: int = 2):
    targets = []
    for row in range(start_row, ws.max_row + 1):
        raw = ws[f"A{row}"].value
        if raw is None or str(raw).strip() == "":
            continue
        targets.append(row)
    return targets
